import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
# Open the Excel file
excel_file = 'output1.xlsx'
workbook = openpyxl.load_workbook(excel_file)
worksheet = workbook.active

# Apply formatting to cells
title_font = Font(bold=True)
price_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
center_alignment = Alignment(horizontal='center')

for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=2):
    # Apply formatting to Title column (column index 1)
    row[0].font = title_font
    row[0].alignment = center_alignment

    # Apply formatting to Price column (column index 2)
    row[1].fill = price_fill
    row[1].alignment = center_alignment

# Save the modified Excel file
output_file = 'output2.xlsx'
workbook.save(output_file)

print(f"Formatting applied and saved to '{output_file}'.")
